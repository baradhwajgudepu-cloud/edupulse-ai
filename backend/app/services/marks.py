import uuid
import logging
from datetime import datetime, timezone
from typing import List, Optional, Dict, Any
from fastapi import HTTPException, status
from sqlalchemy import select
from sqlalchemy.orm import joinedload

from app.models.marks import Marks, MarksStatus, ExamResult
from app.models.teacher import Teacher
from app.models.teacher_subject_assignment import TeacherSubjectAssignment
from app.models.student import Student
from app.models.examination import Examination, ExamSchedule, ExamStatus
from app.models.user import User
from app.repositories.marks import MarksRepository
from app.repositories.examination import ExamScheduleRepository, ExaminationRepository
from app.repositories.student import StudentRepository
from app.repositories.teacher_subject_assignment import TeacherSubjectAssignmentRepository
from app.repositories.school import SchoolRepository
from app.schemas.marks import (
    BulkMarksEntry, SingleMarkEntry,
    SmartMissingSummary, MarkWizardItem, StudentShortInfo,
    PublishSummaryResponse, ResultSummaryResponse, MarksResponse,
    MarksReviewQueueItem, ParentExamResultResponse, ParentSubjectMarkItem,
    ParentTimetableSlot, ParentReportCardItem, MarksExcelUploadSummary,
    ExamWideUploadRowPreview, ExamWideUploadPreviewResponse,
    ExamWideUploadConfirmRequest, ExamWideUploadSummary, ExaminationPublishSummary
)
from app.models.guardian import Guardian, StudentGuardian
from app.services.notification import NotificationService

logger = logging.getLogger(__name__)

class MarksService:
    def __init__(
        self,
        marks_repo: MarksRepository,
        schedule_repo: ExamScheduleRepository,
        exam_repo: ExaminationRepository,
        student_repo: StudentRepository,
        tsa_repo: TeacherSubjectAssignmentRepository,
        school_repo: SchoolRepository,
        notification_service: NotificationService
    ) -> None:
        self.marks_repo = marks_repo
        self.schedule_repo = schedule_repo
        self.exam_repo = exam_repo
        self.student_repo = student_repo
        self.tsa_repo = tsa_repo
        self.school_repo = school_repo
        self.notification_service = notification_service

    def _validate_marks_transition(self, current_status: MarksStatus, target_status: MarksStatus) -> None:
        """
        Canonical Marks Workflow State Machine:
        Allowed Transitions:
          DRAFT -> SUBMITTED
          RETURNED -> SUBMITTED
          SUBMITTED -> UNDER_REVIEW
          SUBMITTED -> APPROVED
          SUBMITTED -> RETURNED
          UNDER_REVIEW -> APPROVED
          UNDER_REVIEW -> RETURNED
          APPROVED -> PUBLISHED
          PUBLISHED -> LOCKED

        Self-transitions are allowed for idempotency.
        All other transitions raise 422 Unprocessable Content.
        """
        if current_status == target_status:
            return

        valid_transitions = {
            MarksStatus.DRAFT: {MarksStatus.SUBMITTED},
            MarksStatus.RETURNED: {MarksStatus.SUBMITTED},
            MarksStatus.SUBMITTED: {MarksStatus.UNDER_REVIEW, MarksStatus.APPROVED, MarksStatus.RETURNED},
            MarksStatus.UNDER_REVIEW: {MarksStatus.APPROVED, MarksStatus.RETURNED},
            MarksStatus.APPROVED: {MarksStatus.PUBLISHED},
            MarksStatus.PUBLISHED: {MarksStatus.LOCKED},
            MarksStatus.LOCKED: set(), # Locked marks are immutable and cannot transition
        }

        allowed = valid_transitions.get(current_status, set())
        if target_status not in allowed:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Illegal workflow transition from {current_status.value if hasattr(current_status, 'value') else current_status} to {target_status.value if hasattr(target_status, 'value') else target_status}."
            )

    async def get_wizard_entry(
        self, tenant_id: uuid.UUID, school_id: uuid.UUID, exam_schedule_id: uuid.UUID, current_user: User
    ) -> SmartMissingSummary:
        # 1. Fetch schedule
        stmt_s = select(ExamSchedule).where(
            ExamSchedule.id == exam_schedule_id,
            ExamSchedule.school_id == school_id,
            ExamSchedule.tenant_id == tenant_id,
            ExamSchedule.deleted_at.is_(None)
        )
        res_s = await self.marks_repo.db.execute(stmt_s)
        sched = res_s.scalar_one_or_none()
        if not sched:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Exam schedule slot not found."
            )

        # 2. Fetch sorted class students list
        students = await self.marks_repo.get_class_students_sorted(sched.class_id, sched.section_id, school_id, tenant_id)

        # 3. Fetch existing mark records
        existing_marks = await self.marks_repo.get_by_schedule_id(exam_schedule_id, tenant_id)
        marks_map = {m.student_id: m for m in existing_marks}

        # 4. Compute statistics
        total = len(students)
        entered_count = 0
        scores = []
        missing_students = []
        wizard_items = []

        for student in students:
            record = marks_map.get(student.id)
            short_info = StudentShortInfo(
                id=student.id,
                first_name=student.first_name,
                last_name=student.last_name,
                roll_number=student.roll_number
            )
            
            if record:
                entered_count += 1
                if record.result_status in [ExamResult.PRESENT, ExamResult.EXEMPTED] and record.marks_obtained is not None:
                    scores.append(float(record.marks_obtained))
                
                wizard_items.append(MarkWizardItem(
                    student=short_info,
                    mark_record=MarksResponse.model_validate(record),
                    is_missing=False
                ))
            else:
                missing_students.append(short_info)
                wizard_items.append(MarkWizardItem(
                    student=short_info,
                    mark_record=None,
                    is_missing=True
                ))

        missing_count = total - entered_count
        avg = round(sum(scores) / len(scores), 2) if scores else None
        high = max(scores) if scores else None
        low = min(scores) if scores else None

        return SmartMissingSummary(
            total_students=total,
            entered_count=entered_count,
            missing_count=missing_count,
            average_score=avg,
            highest_score=high,
            lowest_score=low,
            missing_students=missing_students,
            entries=wizard_items
        )

    async def bulk_save_marks(
        self, tenant_id: uuid.UUID, school_id: uuid.UUID, obj_in: BulkMarksEntry, current_user: User, autosave: bool = False
    ) -> List[Marks]:
        # 1. Fetch schedule
        stmt_s = select(ExamSchedule).where(
            ExamSchedule.id == obj_in.exam_schedule_id,
            ExamSchedule.school_id == school_id,
            ExamSchedule.tenant_id == tenant_id,
            ExamSchedule.deleted_at.is_(None)
        )
        res_s = await self.marks_repo.db.execute(stmt_s)
        sched = res_s.scalar_one_or_none()
        if not sched:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Exam schedule slot not found."
            )

        # 2. Check Lock constraints on Examination status
        stmt_e = select(Examination).where(
            Examination.id == sched.exam_id,
            Examination.deleted_at.is_(None)
        )
        res_e = await self.marks_repo.db.execute(stmt_e)
        exam = res_e.scalar_one_or_none()
        if not exam:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Examination master not found.")

        if exam.status in [ExamStatus.LOCKED, ExamStatus.COMPLETED, ExamStatus.ARCHIVED]:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail="Cannot modify marks because the examination is frozen."
            )

        # 3. TSA Assignment validation (optional for administrative superset operations)
        tsa = None
        if obj_in.teacher_subject_assignment_id:
            tsa = await self.tsa_repo.get_by_id(obj_in.teacher_subject_assignment_id, school_id, tenant_id)
            if not tsa or not tsa.is_active:
                raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Active Teacher Subject Assignment not found.")
            if tsa.class_id != sched.class_id or tsa.section_id != sched.section_id or tsa.subject_id != sched.subject_id:
                raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="TSA assignment parameters do not match exam schedule details.")
        else:
            stmt_auto = select(TeacherSubjectAssignment).where(
                TeacherSubjectAssignment.class_id == sched.class_id,
                TeacherSubjectAssignment.section_id == sched.section_id,
                TeacherSubjectAssignment.subject_id == sched.subject_id,
                TeacherSubjectAssignment.school_id == school_id,
                TeacherSubjectAssignment.tenant_id == tenant_id,
                TeacherSubjectAssignment.deleted_at.is_(None),
                TeacherSubjectAssignment.is_active == True
            )
            res_auto = await self.marks_repo.db.execute(stmt_auto)
            tsa = res_auto.scalars().first()

        effective_tsa_id = tsa.id if tsa else sched.id
        effective_teacher_id = tsa.teacher_id if tsa else current_user.id

        saved_marks = []
        last_saved_student_id = None
        last_saved_roll_number = None

        try:
            # 4. Iterate and validate each marks entry in a single transaction block
            for entry in obj_in.marks:
                # Check student details and containment
                stmt_st = select(Student).where(
                    Student.id == entry.student_id,
                    Student.school_id == school_id,
                    Student.tenant_id == tenant_id,
                    Student.deleted_at.is_(None)
                )
                res_st = await self.marks_repo.db.execute(stmt_st)
                student = res_st.scalar_one_or_none()
                if not student or not student.is_active:
                    raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail=f"Active student not found for ID: {entry.student_id}")

                if student.class_id != sched.class_id or student.section_id != sched.section_id:
                    raise HTTPException(
                        status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                        detail=f"Student {student.first_name} does not belong to target class/section of the scheduled exam."
                    )

                # Mark details sanitization
                marks_obtained = entry.marks_obtained
                if entry.result_status in [ExamResult.ABSENT, ExamResult.MALPRACTICE]:
                    marks_obtained = None
                elif entry.result_status == ExamResult.PRESENT and not autosave:
                    # If it's a save action (not autosave), force check value
                    if marks_obtained is None:
                        raise HTTPException(
                            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                            detail=f"Present status requires a mark value for student {student.first_name}."
                        )
                elif entry.result_status == ExamResult.EXEMPTED:
                    # Exempted students may have None marks
                    pass

                if marks_obtained is not None:
                    if marks_obtained < 0 or marks_obtained > sched.max_marks:
                        raise HTTPException(
                            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                            detail=f"Marks obtained ({marks_obtained}) must be between 0 and maximum marks ({sched.max_marks}) for student {student.first_name}."
                        )

                # Check duplicate / load existing record
                db_mark = await self.marks_repo.get_by_student_and_schedule(sched.id, student.id, tenant_id)

                if db_mark:
                    if db_mark.status == MarksStatus.LOCKED and not current_user.is_superuser and "SUPER_ADMIN" not in [r.code for r in current_user.roles]:
                        raise HTTPException(
                            status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                            detail=f"Marks entries are already locked for student {student.first_name}."
                        )

                    # Optimistic concurrency check
                    if entry.version is not None and db_mark.version != entry.version:
                        raise HTTPException(
                            status_code=status.HTTP_409_CONFLICT,
                            detail={"detail": "Marks were modified by another user. Refresh and try again.", "code": "MARKS_CONFLICT"}
                        )

                    # Record corrections in Audit History JSONB array
                    if db_mark.marks_obtained != marks_obtained or db_mark.result_status != entry.result_status or entry.override_reason:
                        action_type = "MARK_OVERRIDDEN" if entry.override_reason else "MARK_UPDATED"
                        audit_record = {
                            "action": action_type,
                            "reason": entry.override_reason or entry.remarks or "Marks updated",
                            "old_marks": float(db_mark.marks_obtained) if db_mark.marks_obtained is not None else None,
                            "new_marks": float(marks_obtained) if marks_obtained is not None else None,
                            "old_status": db_mark.status.value if hasattr(db_mark.status, "value") else str(db_mark.status),
                            "new_status": (entry.status or db_mark.status).value if hasattr(entry.status or db_mark.status, "value") else str(entry.status or db_mark.status),
                            "updated_by": str(current_user.id),
                            "updated_at": datetime.now(timezone.utc).isoformat(),
                            "tenant_id": str(tenant_id),
                            "school_id": str(school_id)
                        }
                        db_mark.audit_history = list(db_mark.audit_history or []) + [audit_record]

                    db_mark.marks_obtained = marks_obtained
                    db_mark.result_status = entry.result_status
                    db_mark.remarks = entry.remarks
                    if entry.status is not None:
                        db_mark.status = entry.status
                    db_mark.version = db_mark.version + 1
                    db_mark.updated_by = current_user.id
                    self.marks_repo.db.add(db_mark)
                    saved_marks.append(db_mark)
                else:
                    initial_status = entry.status or MarksStatus.DRAFT
                    init_audit = [{
                        "action": "MARK_CREATED",
                        "reason": entry.override_reason or entry.remarks or "Initial entry",
                        "old_marks": None,
                        "new_marks": float(marks_obtained) if marks_obtained is not None else None,
                        "old_status": None,
                        "new_status": initial_status.value if hasattr(initial_status, "value") else str(initial_status),
                        "updated_by": str(current_user.id),
                        "updated_at": datetime.now(timezone.utc).isoformat(),
                        "tenant_id": str(tenant_id),
                        "school_id": str(school_id)
                    }]
                    db_mark = Marks(
                        tenant_id=tenant_id,
                        school_id=school_id,
                        academic_year_id=sched.academic_year_id,
                        examination_id=sched.exam_id,
                        exam_schedule_id=sched.id,
                        student_id=student.id,
                        teacher_subject_assignment_id=effective_tsa_id,
                        teacher_id=effective_teacher_id,
                        subject_id=sched.subject_id,
                        class_id=sched.class_id,
                        section_id=sched.section_id,
                        maximum_marks=sched.max_marks,
                        marks_obtained=marks_obtained,
                        result_status=entry.result_status,
                        remarks=entry.remarks,
                        status=initial_status,
                        version=1,
                        audit_history=init_audit,
                        created_by=current_user.id,
                        updated_by=current_user.id
                    )
                    self.marks_repo.db.add(db_mark)
                    saved_marks.append(db_mark)

                last_saved_student_id = str(student.id)
                last_saved_roll_number = student.roll_number

            # 5. Save the Resume Session pointer state inside TeacherSubjectAssignment settings if TSA exists
            if last_saved_student_id and tsa:
                tsa.settings = {
                    **(tsa.settings or {}),
                    "last_marks_session": {
                        "exam_schedule_id": str(obj_in.exam_schedule_id),
                        "last_student_id": last_saved_student_id,
                        "last_roll_number": last_saved_roll_number
                    }
                }
                self.marks_repo.db.add(tsa)

            await self.marks_repo.db.commit()
        except Exception as e:
            await self.marks_repo.db.rollback()
            raise e
        
        # Flush session and refresh items
        refreshed_marks = []
        for sm in saved_marks:
            refreshed_marks.append(await self.marks_repo.get_by_id(sm.id, school_id, tenant_id))
        return refreshed_marks

    async def get_publish_summary(
        self, tenant_id: uuid.UUID, school_id: uuid.UUID, exam_schedule_id: uuid.UUID
    ) -> PublishSummaryResponse:
        # Fetch schedule
        stmt_s = select(ExamSchedule).where(
            ExamSchedule.id == exam_schedule_id,
            ExamSchedule.school_id == school_id,
            ExamSchedule.tenant_id == tenant_id
        )
        res_s = await self.marks_repo.db.execute(stmt_s)
        sched = res_s.scalar_one_or_none()
        if not sched:
            raise HTTPException(status_code=404, detail="Schedule not found.")

        # Load exam name and subject name
        stmt_e = select(Examination).where(Examination.id == sched.exam_id)
        res_e = await self.marks_repo.db.execute(stmt_e)
        exam = res_e.scalar_one_or_none()

        # Load class name
        students = await self.marks_repo.get_class_students_sorted(sched.class_id, sched.section_id, school_id, tenant_id)
        total = len(students)

        # Load marks records
        marks = await self.marks_repo.get_by_schedule_id(exam_schedule_id, tenant_id)
        entered = len(marks)
        missing = total - entered

        pass_count = 0
        eval_count = 0
        for m in marks:
            if m.result_status in [ExamResult.PRESENT, ExamResult.EXEMPTED]:
                eval_count += 1
                if m.marks_obtained is not None and m.marks_obtained >= sched.pass_marks:
                    pass_count += 1

        pass_pct = round((pass_count / eval_count) * 100, 2) if eval_count else 0.0

        return PublishSummaryResponse(
            exam_name=exam.exam_name if exam else "Exam",
            subject_name="Subject",
            class_name="Class",
            total_students=total,
            entered_count=entered,
            missing_count=missing,
            pass_percentage=pass_pct
        )

    async def publish_marks(
        self, tenant_id: uuid.UUID, school_id: uuid.UUID, exam_schedule_id: uuid.UUID, current_user: User
    ) -> List[Marks]:
        marks = await self.marks_repo.get_by_schedule_id(exam_schedule_id, tenant_id)
        if not marks:
            raise HTTPException(status_code=422, detail="No marks logs found to publish.")

        for m in marks:
            self._validate_marks_transition(m.status, MarksStatus.PUBLISHED)
            old_st = m.status.value if hasattr(m.status, "value") else str(m.status)
            m.status = MarksStatus.PUBLISHED
            audit_record = {
                "action": "MARKS_PUBLISHED",
                "reason": "Marks published for official results",
                "old_status": old_st,
                "new_status": "PUBLISHED",
                "updated_by": str(current_user.id),
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "tenant_id": str(tenant_id),
                "school_id": str(school_id)
            }
            m.audit_history = list(m.audit_history or []) + [audit_record]
            m.updated_by = current_user.id
            self.marks_repo.db.add(m)

        await self.marks_repo.db.commit()

        # Trigger notification
        if marks:
            first_mark = marks[0]
            try:
                await self.notification_service.notify_marks(
                    tenant_id=tenant_id,
                    school_id=school_id,
                    exam_id=first_mark.examination_id,
                    class_id=first_mark.class_id,
                    section_id=first_mark.section_id
                )
            except Exception as ne:
                logger.error(f"Failed to send marks notification: {str(ne)}", exc_info=True)

        return await self.marks_repo.get_by_schedule_id(exam_schedule_id, tenant_id)

    async def get_result_summary(
        self, tenant_id: uuid.UUID, school_id: uuid.UUID, exam_schedule_id: uuid.UUID
    ) -> ResultSummaryResponse:
        # Load marks
        marks = await self.marks_repo.get_by_schedule_id(exam_schedule_id, tenant_id)
        students = await self.marks_repo.get_class_students_sorted(
            # Fetch target class/sec from first mark entry or schedule
            marks[0].class_id if marks else uuid.uuid4(),
            marks[0].section_id if marks else uuid.uuid4(),
            school_id, tenant_id
        )
        total = len(students)
        entered = len(marks)
        missing = total - entered

        scores = []
        pass_count = 0
        absent_count = 0
        
        for m in marks:
            if m.result_status == ExamResult.ABSENT:
                absent_count += 1
            if m.result_status in [ExamResult.PRESENT, ExamResult.EXEMPTED] and m.marks_obtained is not None:
                scores.append(float(m.marks_obtained))
                if m.marks_obtained >= m.maximum_marks * 0.35: # Pass limit 35%
                    pass_count += 1

        avg = round(sum(scores) / len(scores), 2) if scores else 0.0
        pass_pct = round((pass_count / len(scores)) * 100, 2) if scores else 0.0
        high = max(scores) if scores else 0.0
        low = min(scores) if scores else 0.0

        return ResultSummaryResponse(
            class_average=avg,
            pass_percentage=pass_pct,
            highest_score=high,
            lowest_score=low,
            missing_count=missing,
            absent_count=absent_count
        )

    async def lock_marks(
        self, tenant_id: uuid.UUID, school_id: uuid.UUID, exam_schedule_id: uuid.UUID, current_user: User
    ) -> List[Marks]:
        marks = await self.marks_repo.get_by_schedule_id(exam_schedule_id, tenant_id)
        if not marks:
            raise HTTPException(status_code=404, detail="No marks records found to lock.")
        for m in marks:
            self._validate_marks_transition(m.status, MarksStatus.LOCKED)
            old_st = m.status.value if hasattr(m.status, "value") else str(m.status)
            m.status = MarksStatus.LOCKED
            audit_record = {
                "action": "MARKS_LOCKED",
                "reason": "Administrative marks locking",
                "old_status": old_st,
                "new_status": "LOCKED",
                "updated_by": str(current_user.id),
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "tenant_id": str(tenant_id),
                "school_id": str(school_id)
            }
            m.audit_history = list(m.audit_history or []) + [audit_record]
            m.updated_by = current_user.id
            self.marks_repo.db.add(m)
        await self.marks_repo.db.commit()
        return await self.marks_repo.get_by_schedule_id(exam_schedule_id, tenant_id)

    async def unlock_marks(
        self, tenant_id: uuid.UUID, school_id: uuid.UUID, exam_schedule_id: uuid.UUID, reason: str, current_user: User
    ) -> List[Marks]:
        if not reason or not reason.strip():
            raise HTTPException(status_code=422, detail="An administrative unlock reason is mandatory.")
        marks = await self.marks_repo.get_by_schedule_id(exam_schedule_id, tenant_id)
        if not marks:
            raise HTTPException(status_code=404, detail="No marks records found to unlock.")
        for m in marks:
            old_st = m.status.value if hasattr(m.status, "value") else str(m.status)
            m.status = MarksStatus.DRAFT
            audit_record = {
                "action": "MARK_OVERRIDDEN",
                "reason": reason.strip(),
                "old_status": old_st,
                "new_status": "DRAFT",
                "updated_by": str(current_user.id),
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "tenant_id": str(tenant_id),
                "school_id": str(school_id)
            }
            m.audit_history = list(m.audit_history or []) + [audit_record]
            m.updated_by = current_user.id
            self.marks_repo.db.add(m)
        await self.marks_repo.db.commit()
        return await self.marks_repo.get_by_schedule_id(exam_schedule_id, tenant_id)

    async def submit_for_review(
        self, tenant_id: uuid.UUID, school_id: uuid.UUID, exam_schedule_id: uuid.UUID, notes: Optional[str], current_user: User, allow_partial: bool = False
    ) -> List[Marks]:
        stmt_s = select(ExamSchedule).where(
            ExamSchedule.id == exam_schedule_id,
            ExamSchedule.school_id == school_id,
            ExamSchedule.tenant_id == tenant_id
        )
        res_s = await self.marks_repo.db.execute(stmt_s)
        sched = res_s.scalar_one_or_none()
        if not sched:
            raise HTTPException(status_code=404, detail="Schedule not found.")

        # Validate all active enrolled students in the section
        enrolled_students = await self.marks_repo.get_class_students_sorted(sched.class_id, sched.section_id, school_id, tenant_id)
        marks = await self.marks_repo.get_by_schedule_id(exam_schedule_id, tenant_id)
        if not marks:
            raise HTTPException(status_code=400, detail="No marks records found to submit for review.")

        marked_student_ids = {m.student_id for m in marks}
        missing_students = [s for s in enrolled_students if s.id not in marked_student_ids]

        if missing_students and not allow_partial:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"Cannot submit marks. {len(missing_students)} enrolled student(s) have missing marks entries. Complete all entries or submit with explicit partial submission flag."
            )

        for m in marks:
            self._validate_marks_transition(m.status, MarksStatus.SUBMITTED)
            old_st = m.status.value if hasattr(m.status, "value") else str(m.status)
            m.status = MarksStatus.SUBMITTED
            audit_record = {
                "action": "MARKS_SUBMITTED" if not missing_students else "PARTIAL_MARKS_SUBMITTED",
                "notes": notes or ("Partial teacher marks submission" if missing_students else "Teacher submitted marks for Principal review"),
                "reason": notes or "Teacher submission",
                "old_status": old_st,
                "new_status": "SUBMITTED",
                "updated_by": str(current_user.id),
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "tenant_id": str(tenant_id),
                "school_id": str(school_id)
            }
            m.audit_history = list(m.audit_history or []) + [audit_record]
            m.updated_by = current_user.id
            self.marks_repo.db.add(m)
        
        await self.marks_repo.db.commit()
        return await self.marks_repo.get_by_schedule_id(exam_schedule_id, tenant_id)

    async def approve_marks(
        self, tenant_id: uuid.UUID, school_id: uuid.UUID, exam_schedule_id: uuid.UUID, remarks: Optional[str], current_user: User
    ) -> List[Marks]:
        marks = await self.marks_repo.get_by_schedule_id(exam_schedule_id, tenant_id)
        if not marks:
            raise HTTPException(status_code=404, detail="No marks records found to approve.")
        
        for m in marks:
            self._validate_marks_transition(m.status, MarksStatus.APPROVED)
            old_st = m.status.value if hasattr(m.status, "value") else str(m.status)
            m.status = MarksStatus.APPROVED
            audit_record = {
                "action": "MARKS_APPROVED",
                "remarks": remarks or "Approved by Principal",
                "reason": remarks or "Principal approval",
                "old_status": old_st,
                "new_status": "APPROVED",
                "updated_by": str(current_user.id),
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "tenant_id": str(tenant_id),
                "school_id": str(school_id)
            }
            m.audit_history = list(m.audit_history or []) + [audit_record]
            m.updated_by = current_user.id
            self.marks_repo.db.add(m)
        
        await self.marks_repo.db.commit()
        return await self.marks_repo.get_by_schedule_id(exam_schedule_id, tenant_id)

    async def return_marks_for_correction(
        self, tenant_id: uuid.UUID, school_id: uuid.UUID, exam_schedule_id: uuid.UUID, correction_reason: str, current_user: User
    ) -> List[Marks]:
        if not correction_reason or not correction_reason.strip():
            raise HTTPException(status_code=422, detail="A mandatory correction reason is required when returning marks.")
        
        marks = await self.marks_repo.get_by_schedule_id(exam_schedule_id, tenant_id)
        if not marks:
            raise HTTPException(status_code=404, detail="No marks records found to return.")
        
        for m in marks:
            self._validate_marks_transition(m.status, MarksStatus.RETURNED)
            old_st = m.status.value if hasattr(m.status, "value") else str(m.status)
            m.status = MarksStatus.RETURNED
            settings_dict = dict(m.settings or {})
            settings_dict["correction_reason"] = correction_reason.strip()
            m.settings = settings_dict
            audit_record = {
                "action": "MARKS_RETURNED",
                "reason": correction_reason.strip(),
                "old_status": old_st,
                "new_status": "RETURNED",
                "updated_by": str(current_user.id),
                "updated_at": datetime.now(timezone.utc).isoformat(),
                "tenant_id": str(tenant_id),
                "school_id": str(school_id)
            }
            m.audit_history = list(m.audit_history or []) + [audit_record]
            m.updated_by = current_user.id
            self.marks_repo.db.add(m)
        
        await self.marks_repo.db.commit()
        return await self.marks_repo.get_by_schedule_id(exam_schedule_id, tenant_id)

    async def get_review_queue(
        self, tenant_id: uuid.UUID, school_id: uuid.UUID, examination_id: Optional[uuid.UUID] = None, academic_year_id: Optional[uuid.UUID] = None
    ) -> List[MarksReviewQueueItem]:
        from sqlalchemy.orm import joinedload

        stmt = select(ExamSchedule).where(
            ExamSchedule.school_id == school_id,
            ExamSchedule.tenant_id == tenant_id,
            ExamSchedule.deleted_at.is_(None)
        ).options(
            joinedload(ExamSchedule.examination),
            joinedload(ExamSchedule.class_obj),
            joinedload(ExamSchedule.section),
            joinedload(ExamSchedule.subject)
        )
        if examination_id:
            stmt = stmt.where(ExamSchedule.exam_id == examination_id)
        if academic_year_id:
            stmt = stmt.where(ExamSchedule.academic_year_id == academic_year_id)
        
        stmt = stmt.order_by(ExamSchedule.exam_date.asc())
        res = await self.marks_repo.db.execute(stmt)
        schedules = list(res.unique().scalars().all())
        
        queue_items = []
        for sched in schedules:
            students = await self.marks_repo.get_class_students_sorted(sched.class_id, sched.section_id, school_id, tenant_id)
            marks = await self.marks_repo.get_by_schedule_id(sched.id, tenant_id)
            total = len(students)
            entered = len(marks)
            missing = total - entered
            
            scores = []
            pass_count = 0
            fail_count = 0
            absent_count = 0
            correction_reason = None
            last_submitted_at = None
            
            batch_status = MarksStatus.DRAFT
            if marks:
                statuses = [m.status for m in marks]
                if all(s == MarksStatus.PUBLISHED for s in statuses):
                    batch_status = MarksStatus.PUBLISHED
                elif all(s == MarksStatus.LOCKED for s in statuses):
                    batch_status = MarksStatus.LOCKED
                elif any(s == MarksStatus.RETURNED for s in statuses):
                    batch_status = MarksStatus.RETURNED
                elif all(s == MarksStatus.APPROVED for s in statuses):
                    batch_status = MarksStatus.APPROVED
                elif any(s in [MarksStatus.SUBMITTED, MarksStatus.UNDER_REVIEW] for s in statuses):
                    batch_status = MarksStatus.SUBMITTED
                else:
                    batch_status = MarksStatus.DRAFT
                
                for m in marks:
                    if m.result_status == ExamResult.ABSENT:
                        absent_count += 1
                    if m.result_status in [ExamResult.PRESENT, ExamResult.EXEMPTED] and m.marks_obtained is not None:
                        val = float(m.marks_obtained)
                        scores.append(val)
                        if val >= sched.pass_marks:
                            pass_count += 1
                        else:
                            fail_count += 1
                    if m.settings and "correction_reason" in m.settings:
                        correction_reason = m.settings["correction_reason"]
                    if m.audit_history:
                        for h in reversed(m.audit_history):
                            if h.get("action") == "SUBMIT_FOR_REVIEW" and "updated_at" in h:
                                try:
                                    last_submitted_at = datetime.fromisoformat(h["updated_at"])
                                except Exception:
                                    pass
                                break
            
            avg = round(sum(scores) / len(scores), 2) if scores else None
            pass_pct = round((pass_count / len(scores)) * 100, 2) if scores else 0.0
            high = max(scores) if scores else None
            low = min(scores) if scores else None
            
            # Fetch assigned teacher name
            teacher_name = None
            teacher_id = None
            stmt_tsa = select(TeacherSubjectAssignment).where(
                TeacherSubjectAssignment.class_id == sched.class_id,
                TeacherSubjectAssignment.section_id == sched.section_id,
                TeacherSubjectAssignment.subject_id == sched.subject_id,
                TeacherSubjectAssignment.tenant_id == tenant_id
            ).options(joinedload(TeacherSubjectAssignment.teacher))
            res_tsa = await self.marks_repo.db.execute(stmt_tsa)
            tsa = res_tsa.scalar_one_or_none()
            if tsa and tsa.teacher:
                teacher_id = tsa.teacher.id
                teacher_name = f"{tsa.teacher.first_name} {tsa.teacher.last_name}".strip()
            
            queue_items.append(MarksReviewQueueItem(
                exam_schedule_id=sched.id,
                examination_id=sched.exam_id,
                exam_name=getattr(sched.examination, "exam_name", getattr(sched.examination, "name", "Exam")) if sched.examination else "Exam",
                class_id=sched.class_id,
                class_name=sched.class_obj.name if sched.class_obj else "Class",
                section_id=sched.section_id,
                section_name=sched.section.name if sched.section else "Section",
                subject_id=sched.subject_id,
                subject_name=getattr(sched.subject, "subject_name", getattr(sched.subject, "name", "Subject")) if sched.subject else "Subject",
                teacher_id=teacher_id,
                teacher_name=teacher_name,
                exam_date=str(sched.exam_date),
                max_marks=sched.max_marks,
                pass_marks=sched.pass_marks,
                total_students=total,
                entered_count=entered,
                missing_count=missing,
                pass_count=pass_count,
                fail_count=fail_count,
                absent_count=absent_count,
                average_score=avg,
                highest_score=high,
                lowest_score=low,
                pass_percentage=pass_pct,
                batch_status=batch_status,
                last_submitted_at=last_submitted_at,
                correction_reason=correction_reason
            ))
        return queue_items

    async def get_parent_student_marks(
        self, tenant_id: uuid.UUID, school_id: uuid.UUID, student_id: uuid.UUID, current_user: User
    ) -> List[ParentExamResultResponse]:
        from sqlalchemy.orm import joinedload
        from sqlalchemy import or_

        role_codes = {r.code for r in current_user.roles}
        is_admin = current_user.is_superuser or any(code in ["SUPER_ADMIN", "SCHOOL_ADMIN", "PRINCIPAL"] for code in role_codes)
        if not is_admin:
            stmt_guard = select(StudentGuardian).join(Guardian).where(
                StudentGuardian.student_id == student_id,
                StudentGuardian.school_id == school_id,
                StudentGuardian.tenant_id == tenant_id,
                or_(
                    Guardian.user_id == current_user.id,
                    Guardian.email == current_user.email
                )
            )
            res_guard = await self.marks_repo.db.execute(stmt_guard)
            if not res_guard.scalar_one_or_none():
                raise HTTPException(status_code=403, detail="Access denied. You are not a linked guardian of this student.")

        stmt_st = select(Student).where(
            Student.id == student_id,
            Student.school_id == school_id,
            Student.tenant_id == tenant_id
        ).options(
            joinedload(Student.class_obj),
            joinedload(Student.section)
        )
        res_st = await self.marks_repo.db.execute(stmt_st)
        student = res_st.scalar_one_or_none()
        if not student:
            raise HTTPException(status_code=404, detail="Student not found.")

        stmt_m = select(Marks).where(
            Marks.student_id == student_id,
            Marks.school_id == school_id,
            Marks.tenant_id == tenant_id,
            Marks.status == MarksStatus.PUBLISHED,
            Marks.deleted_at.is_(None)
        ).options(
            joinedload(Marks.examination).joinedload(Examination.academic_year),
            joinedload(Marks.subject)
        ).order_by(Marks.created_at.asc())
        
        res_m = await self.marks_repo.db.execute(stmt_m)
        published_marks = list(res_m.unique().scalars().all())

        exams_map: Dict[uuid.UUID, List[Marks]] = {}
        for m in published_marks:
            exams_map.setdefault(m.examination_id, []).append(m)

        responses = []
        for exam_id, m_list in exams_map.items():
            first = m_list[0]
            exam = first.examination
            ay_name = exam.academic_year.name if exam and exam.academic_year else "Current"
            
            sub_items = []
            total_max = 0
            total_obt = 0.0
            all_passed = True
            
            for m in m_list:
                obt = float(m.marks_obtained) if m.marks_obtained is not None else 0.0
                if m.result_status == ExamResult.PRESENT:
                    total_max += m.maximum_marks
                    total_obt += obt
                    is_pass = (m.marks_obtained or 0) >= (m.maximum_marks * 0.35)
                    if not is_pass:
                        all_passed = False
                elif m.result_status == ExamResult.EXEMPTED:
                    is_pass = True
                else: # ABSENT, MALPRACTICE, WITHHELD
                    total_max += m.maximum_marks
                    is_pass = False
                    all_passed = False
                
                sub_items.append(ParentSubjectMarkItem(
                    subject_id=m.subject_id,
                    subject_name=getattr(m.subject, "subject_name", getattr(m.subject, "name", "Subject")) if m.subject else "Subject",
                    subject_code=getattr(m.subject, "subject_code", getattr(m.subject, "code", None)) if m.subject else None,
                    exam_date=None,
                    maximum_marks=m.maximum_marks,
                    pass_marks=int(m.maximum_marks * 0.35),
                    marks_obtained=float(m.marks_obtained) if m.marks_obtained is not None else None,
                    result_status=m.result_status,
                    grade=m.grade,
                    remarks=m.remarks,
                    is_passed=is_pass
                ))
            
            pct = round((total_obt / total_max) * 100, 2) if total_max > 0 else 0.0
            status_str = "PASSED" if all_passed else "FAILED"
            
            exam_name_val = getattr(exam, "exam_name", getattr(exam, "name", "Examination")) if exam else "Examination"
            exam_type_val = exam.exam_type.value if exam and hasattr(exam.exam_type, "value") else (str(exam.exam_type) if exam else "SCHOLASTIC")

            responses.append(ParentExamResultResponse(
                examination_id=exam_id,
                exam_name=exam_name_val,
                exam_type=exam_type_val,
                academic_year_name=ay_name,
                student_id=student.id,
                student_name=f"{student.first_name} {student.last_name}".strip(),
                roll_number=student.roll_number,
                class_name=student.class_obj.name if student.class_obj else "Class",
                section_name=student.section.name if student.section else "Section",
                total_max_marks=total_max,
                total_obtained_marks=round(total_obt, 2),
                overall_percentage=pct,
                overall_grade="A+" if pct >= 90 else "A" if pct >= 80 else "B" if pct >= 70 else "C" if pct >= 60 else "D" if pct >= 50 else "F",
                status=status_str,
                subject_marks=sub_items
            ))
        return responses

    async def get_parent_student_timetable(
        self, tenant_id: uuid.UUID, school_id: uuid.UUID, student_id: uuid.UUID, current_user: User
    ) -> List[ParentTimetableSlot]:
        from sqlalchemy.orm import joinedload
        from sqlalchemy import or_

        role_codes = {r.code for r in current_user.roles}
        is_admin = current_user.is_superuser or any(code in ["SUPER_ADMIN", "SCHOOL_ADMIN", "PRINCIPAL"] for code in role_codes)
        if not is_admin:
            stmt_guard = select(StudentGuardian).join(Guardian).where(
                StudentGuardian.student_id == student_id,
                StudentGuardian.school_id == school_id,
                StudentGuardian.tenant_id == tenant_id,
                or_(
                    Guardian.user_id == current_user.id,
                    Guardian.email == current_user.email
                )
            )
            res_guard = await self.marks_repo.db.execute(stmt_guard)
            if not res_guard.scalar_one_or_none():
                raise HTTPException(status_code=403, detail="Access denied. You are not a linked guardian of this student.")

        stmt_st = select(Student).where(
            Student.id == student_id,
            Student.school_id == school_id,
            Student.tenant_id == tenant_id
        )
        res_st = await self.marks_repo.db.execute(stmt_st)
        student = res_st.scalar_one_or_none()
        if not student:
            raise HTTPException(status_code=404, detail="Student not found.")

        stmt_s = select(ExamSchedule).where(
            ExamSchedule.class_id == student.class_id,
            ExamSchedule.section_id == student.section_id,
            ExamSchedule.school_id == school_id,
            ExamSchedule.tenant_id == tenant_id,
            ExamSchedule.deleted_at.is_(None)
        ).options(
            joinedload(ExamSchedule.examination),
            joinedload(ExamSchedule.subject)
        ).order_by(ExamSchedule.exam_date.asc(), ExamSchedule.start_time.asc())

        res_s = await self.marks_repo.db.execute(stmt_s)
        schedules = list(res_s.unique().scalars().all())

        return [
            ParentTimetableSlot(
                exam_schedule_id=s.id,
                examination_id=s.exam_id,
                exam_name=getattr(s.examination, "exam_name", getattr(s.examination, "name", "Examination")) if s.examination else "Examination",
                exam_type=s.examination.exam_type.value if (s.examination and hasattr(s.examination.exam_type, "value")) else (str(s.examination.exam_type) if s.examination else "SCHOLASTIC"),
                subject_id=s.subject_id,
                subject_name=getattr(s.subject, "subject_name", getattr(s.subject, "name", "Subject")) if s.subject else "Subject",
                exam_date=str(s.exam_date),
                start_time=s.start_time.strftime("%H:%M"),
                end_time=s.end_time.strftime("%H:%M"),
                max_marks=s.max_marks,
                pass_marks=s.pass_marks,
                room_number=s.room_number
            )
            for s in schedules
        ]

    async def get_parent_student_report_cards(
        self, tenant_id: uuid.UUID, school_id: uuid.UUID, student_id: uuid.UUID, current_user: User
    ) -> List[ParentReportCardItem]:
        from sqlalchemy.orm import joinedload
        from sqlalchemy import or_
        from app.models.report_card import ReportCardPublication, ReportCardStatus

        role_codes = {r.code for r in current_user.roles}
        is_admin = current_user.is_superuser or any(code in ["SUPER_ADMIN", "SCHOOL_ADMIN", "PRINCIPAL"] for code in role_codes)
        if not is_admin:
            stmt_guard = select(StudentGuardian).join(Guardian).where(
                StudentGuardian.student_id == student_id,
                StudentGuardian.school_id == school_id,
                StudentGuardian.tenant_id == tenant_id,
                or_(
                    Guardian.user_id == current_user.id,
                    Guardian.email == current_user.email
                )
            )
            res_guard = await self.marks_repo.db.execute(stmt_guard)
            if not res_guard.scalar_one_or_none():
                raise HTTPException(status_code=403, detail="Access denied. You are not a linked guardian of this student.")

        stmt_rc = select(ReportCardPublication).where(
            ReportCardPublication.student_id == student_id,
            ReportCardPublication.school_id == school_id,
            ReportCardPublication.tenant_id == tenant_id,
            ReportCardPublication.status == ReportCardStatus.PUBLISHED,
            ReportCardPublication.deleted_at.is_(None)
        ).options(
            joinedload(ReportCardPublication.academic_year)
        ).order_by(ReportCardPublication.created_at.desc())

        res_rc = await self.marks_repo.db.execute(stmt_rc)
        report_cards = list(res_rc.unique().scalars().all())

        return [
            ParentReportCardItem(
                report_card_id=rc.id,
                examination_id=uuid.UUID((rc.settings or {}).get("examination_id")) if (rc.settings and (rc.settings or {}).get("examination_id")) else uuid.UUID(int=0),
                exam_name=(rc.settings or {}).get("exam_name", "Academic Progress Report"),
                academic_year_name=rc.academic_year.name if rc.academic_year else "Academic Year",
                status=rc.status.value if hasattr(rc.status, "value") else str(rc.status),
                total_marks=float((rc.settings or {}).get("total_marks", 0.0)),
                percentage=float((rc.settings or {}).get("percentage", 0.0)),
                grade=(rc.settings or {}).get("grade"),
                rank=(rc.settings or {}).get("rank"),
                generated_at=rc.generated_at or rc.created_at,
                pdf_download_url=rc.pdf_url or f"/api/v1/report-cards/{rc.id}/pdf?school_id={school_id}"
            )
            for rc in report_cards
        ]

    async def generate_marks_template(
        self,
        tenant_id: uuid.UUID,
        school_id: uuid.UUID,
        exam_schedule_id: uuid.UUID,
        file_format: str = "xlsx"
    ) -> tuple[bytes, str, str]:
        import io
        import csv
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Fetch schedule and details
        stmt_s = select(ExamSchedule).where(
            ExamSchedule.id == exam_schedule_id,
            ExamSchedule.school_id == school_id,
            ExamSchedule.tenant_id == tenant_id,
            ExamSchedule.deleted_at.is_(None)
        ).options(
            joinedload(ExamSchedule.examination),
            joinedload(ExamSchedule.subject),
            joinedload(ExamSchedule.class_obj),
            joinedload(ExamSchedule.section)
        )
        res_s = await self.marks_repo.db.execute(stmt_s)
        sched = res_s.scalar_one_or_none()
        if not sched:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Exam schedule slot not found.")

        students = await self.marks_repo.get_class_students_sorted(sched.class_id, sched.section_id, school_id, tenant_id)
        existing_marks = await self.marks_repo.get_by_schedule_id(exam_schedule_id, tenant_id)
        marks_map = {m.student_id: m for m in existing_marks}

        subject_name = sched.subject.name if getattr(sched, "subject", None) else "Subject"
        class_name = sched.class_obj.name if getattr(sched, "class_obj", None) else "Class"
        section_name = f"-{sched.section.name}" if getattr(sched, "section", None) else ""
        clean_filename_base = f"marks_template_{subject_name}_{class_name}{section_name}".replace(" ", "_")

        max_marks_val = getattr(sched, "maximum_marks", getattr(sched, "max_marks", 100))

        if file_format.lower() == "csv":
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(["Roll Number", "Student Name", "Max Marks", "Marks Obtained", "Status", "Remarks"])
            for s in students:
                existing = marks_map.get(s.id)
                obtained = existing.marks_obtained if (existing and existing.marks_obtained is not None) else ""
                st_val = existing.result_status.value if (existing and existing.result_status) else "PRESENT"
                rem = existing.remarks if (existing and existing.remarks) else ""
                writer.writerow([s.roll_number or "", s.full_name, max_marks_val, obtained, st_val, rem])
            
            content = output.getvalue().encode("utf-8-sig")
            return content, f"{clean_filename_base}.csv", "text/csv"

        # Default XLSX with openpyxl
        wb = Workbook()
        ws = wb.active
        ws.title = "Marks Entry"

        # Headers and Styling
        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin", color="CCCCCC"),
            right=Side(style="thin", color="CCCCCC"),
            top=Side(style="thin", color="CCCCCC"),
            bottom=Side(style="thin", color="CCCCCC")
        )
        center_align = Alignment(horizontal="center", vertical="center")
        left_align = Alignment(horizontal="left", vertical="center")

        headers = ["Roll Number", "Student Name", "Max Marks", "Marks Obtained", "Status (PRESENT/ABSENT/EXEMPTED/MALPRACTICE)", "Remarks"]
        ws.append(headers)

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = thin_border

        for row_idx, s in enumerate(students, 2):
            existing = marks_map.get(s.id)
            obtained = existing.marks_obtained if (existing and existing.marks_obtained is not None) else None
            st_val = existing.result_status.value if (existing and existing.result_status) else "PRESENT"
            rem = existing.remarks if (existing and existing.remarks) else None

            ws.cell(row=row_idx, column=1, value=s.roll_number or "").alignment = center_align
            ws.cell(row=row_idx, column=2, value=s.full_name).alignment = left_align
            ws.cell(row=row_idx, column=3, value=max_marks_val).alignment = center_align
            ws.cell(row=row_idx, column=4, value=obtained).alignment = center_align
            ws.cell(row=row_idx, column=5, value=st_val).alignment = center_align
            ws.cell(row=row_idx, column=6, value=rem).alignment = left_align

            for c in range(1, 7):
                ws.cell(row=row_idx, column=c).border = thin_border

        # Adjust column widths
        ws.column_dimensions["A"].width = 15
        ws.column_dimensions["B"].width = 25
        ws.column_dimensions["C"].width = 14
        ws.column_dimensions["D"].width = 16
        ws.column_dimensions["E"].width = 45
        ws.column_dimensions["F"].width = 30

        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output.getvalue(), f"{clean_filename_base}.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"

    async def import_marks_from_file(
        self,
        tenant_id: uuid.UUID,
        school_id: uuid.UUID,
        exam_schedule_id: uuid.UUID,
        file_bytes: bytes,
        filename: str,
        current_user: User
    ) -> MarksExcelUploadSummary:
        import io
        import csv
        from openpyxl import load_workbook

        # 1. Fetch schedule
        stmt_s = select(ExamSchedule).where(
            ExamSchedule.id == exam_schedule_id,
            ExamSchedule.school_id == school_id,
            ExamSchedule.tenant_id == tenant_id,
            ExamSchedule.deleted_at.is_(None)
        ).options(
            joinedload(ExamSchedule.examination)
        )
        res_s = await self.marks_repo.db.execute(stmt_s)
        sched = res_s.scalar_one_or_none()
        if not sched:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Exam schedule slot not found.")

        # Check if examination is frozen/archived
        if sched.examination and sched.examination.status in [ExamStatus.ARCHIVED, ExamStatus.COMPLETED]:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="Cannot upload marks for an archived or completed examination.")

        max_marks_limit = getattr(sched, "maximum_marks", getattr(sched, "max_marks", 100))

        # 2. Fetch enrolled students
        students = await self.marks_repo.get_class_students_sorted(sched.class_id, sched.section_id, school_id, tenant_id)
        if not students:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="No enrolled students found for this class and section.")

        # Build lookup indices
        student_by_roll = {str(s.roll_number).strip().lower(): s for s in students if s.roll_number}
        student_by_id = {str(s.id): s for s in students}
        student_by_name = {str(s.full_name).strip().lower(): s for s in students if s.full_name}

        # 3. Parse rows from file
        parsed_rows: List[Dict[str, Any]] = []
        is_csv = filename.lower().endswith(".csv")

        if is_csv:
            try:
                text_content = file_bytes.decode("utf-8-sig", errors="replace")
                reader = csv.reader(io.StringIO(text_content))
                raw_rows = list(reader)
            except Exception as e:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Failed to parse CSV file: {str(e)}")
        else:
            try:
                wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
                ws = wb.active
                raw_rows = [[cell.value for cell in row] for row in ws.iter_rows()]
            except Exception as e:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Failed to parse Excel workbook: {str(e)}")

        if not raw_rows or len(raw_rows) < 2:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File is empty or missing data rows.")

        # Find header positions
        headers = [str(h or "").strip().lower() for h in raw_rows[0]]
        roll_idx = -1
        name_idx = -1
        marks_idx = -1
        status_idx = -1
        remarks_idx = -1

        for idx, h in enumerate(headers):
            if "remark" in h or "comment" in h or "note" in h:
                remarks_idx = idx
            elif "status" in h or "result" in h or "attendance" in h:
                status_idx = idx
            elif "roll" in h or "admission" in h:
                roll_idx = idx
            elif "name" in h or "student" in h:
                name_idx = idx
            elif "obtained" in h or "score" in h or ("mark" in h and "max" not in h):
                marks_idx = idx

        if roll_idx == -1 and name_idx == -1:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Could not detect student identification column ('Roll Number' or 'Student Name') in header."
            )
        if marks_idx == -1:
            marks_idx = 3 if len(headers) > 3 else -1

        errors: List[str] = []
        matched_entries: List[SingleMarkEntry] = []
        total_data_rows = 0

        for row_num, row in enumerate(raw_rows[1:], start=2):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue
            total_data_rows += 1

            roll_val = str(row[roll_idx]).strip() if (roll_idx != -1 and roll_idx < len(row) and row[roll_idx] is not None) else ""
            name_val = str(row[name_idx]).strip() if (name_idx != -1 and name_idx < len(row) and row[name_idx] is not None) else ""
            raw_marks = row[marks_idx] if (marks_idx != -1 and marks_idx < len(row)) else None
            raw_status = str(row[status_idx]).strip().upper() if (status_idx != -1 and status_idx < len(row) and row[status_idx] is not None) else ""
            raw_remarks = str(row[remarks_idx]).strip() if (remarks_idx != -1 and remarks_idx < len(row) and row[remarks_idx] is not None) else None

            # Match student
            matched_student = None
            if roll_val and roll_val.lower() in student_by_roll:
                matched_student = student_by_roll[roll_val.lower()]
            elif roll_val in student_by_id:
                matched_student = student_by_id[roll_val]
            elif name_val and name_val.lower() in student_by_name:
                matched_student = student_by_name[name_val.lower()]

            if not matched_student:
                errors.append(f"Row {row_num}: Student '{roll_val or name_val}' not found in enrolled students.")
                continue

            # Parse status
            result_status = ExamResult.PRESENT
            if "ABSENT" in raw_status or raw_status == "A":
                result_status = ExamResult.ABSENT
            elif "EXEMPT" in raw_status or raw_status == "E":
                result_status = ExamResult.EXEMPTED
            elif "MALPRACTICE" in raw_status or raw_status == "M":
                result_status = ExamResult.MALPRACTICE
            elif raw_status in [e.value for e in ExamResult]:
                result_status = ExamResult(raw_status)

            # Parse marks
            marks_val: Optional[float] = None
            if result_status != ExamResult.ABSENT and raw_marks is not None and str(raw_marks).strip() != "":
                try:
                    marks_val = float(str(raw_marks).strip())
                    if marks_val < 0:
                        errors.append(f"Row {row_num} ({matched_student.full_name}): Marks cannot be negative ({marks_val}).")
                        continue
                    if marks_val > max_marks_limit:
                        errors.append(f"Row {row_num} ({matched_student.full_name}): Marks ({marks_val}) exceed Maximum Marks ({max_marks_limit}).")
                        continue
                except ValueError:
                    errors.append(f"Row {row_num} ({matched_student.full_name}): Invalid marks value '{raw_marks}'.")
                    continue
            elif result_status == ExamResult.ABSENT:
                marks_val = None

            matched_entries.append(
                SingleMarkEntry(
                    student_id=matched_student.id,
                    marks_obtained=marks_val,
                    result_status=result_status,
                    remarks=raw_remarks
                )
            )

        if not matched_entries:
            raise HTTPException(
                status_code=status.HTTP_422_UNPROCESSABLE_ENTITY,
                detail=f"No valid student marks could be processed from file. {len(errors)} error(s) found: {'; '.join(errors[:3])}"
            )

        # 4. Save bulk marks via bulk_save_marks
        bulk_payload = BulkMarksEntry(
            exam_schedule_id=exam_schedule_id,
            marks=matched_entries
        )
        saved_objs = await self.bulk_save_marks(tenant_id, school_id, bulk_payload, current_user, autosave=False)

        return MarksExcelUploadSummary(
            total_rows=total_data_rows,
            matched_students=len(matched_entries),
            saved_count=len(saved_objs),
            errors=errors,
            marks=[MarksResponse.model_validate(m) for m in saved_objs]
        )

    def _compute_grade(self, marks_obtained: Optional[float], max_marks: int) -> Optional[str]:
        if marks_obtained is None or max_marks <= 0:
            return None
        pct = (marks_obtained / max_marks) * 100.0
        if pct >= 90: return "A+"
        elif pct >= 80: return "A"
        elif pct >= 70: return "B+"
        elif pct >= 60: return "B"
        elif pct >= 50: return "C"
        elif pct >= 35: return "D"
        else: return "F"

    async def preview_exam_wide_marks_file(
        self,
        tenant_id: uuid.UUID,
        school_id: uuid.UUID,
        exam_id: uuid.UUID,
        file_bytes: bytes,
        filename: str,
        current_user: User
    ) -> ExamWideUploadPreviewResponse:
        import io
        import csv
        from openpyxl import load_workbook

        # 1. Fetch Examination
        stmt_e = select(Examination).where(
            Examination.id == exam_id,
            Examination.school_id == school_id,
            Examination.tenant_id == tenant_id,
            Examination.deleted_at.is_(None)
        )
        res_e = await self.marks_repo.db.execute(stmt_e)
        examination = res_e.scalar_one_or_none()
        if not examination:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Examination not found.")

        # 2. Fetch all ExamSchedules for this Examination with Class, Section, Subject
        stmt_s = select(ExamSchedule).where(
            ExamSchedule.exam_id == exam_id,
            ExamSchedule.school_id == school_id,
            ExamSchedule.tenant_id == tenant_id,
            ExamSchedule.deleted_at.is_(None)
        ).options(
            joinedload(ExamSchedule.class_obj),
            joinedload(ExamSchedule.section),
            joinedload(ExamSchedule.subject)
        )
        res_s = await self.marks_repo.db.execute(stmt_s)
        schedules = list(res_s.unique().scalars().all())
        if not schedules:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="No exam schedules/papers configured for this examination.")

        # Build Schedule lookup maps: (normalized_class, normalized_section, normalized_subject) -> ExamSchedule
        schedule_map = {}
        for s in schedules:
            c_name = str(s.class_obj.name).strip().lower() if s.class_obj else ""
            c_code = str(s.class_obj.code).strip().lower() if s.class_obj and s.class_obj.code else ""
            sec_name = str(s.section.name).strip().lower() if s.section else ""
            sec_code = str(s.section.code).strip().lower() if s.section and s.section.code else ""
            sub_name = str(s.subject.subject_name).strip().lower() if s.subject else ""
            sub_code = str(s.subject.subject_code).strip().lower() if s.subject and s.subject.subject_code else ""

            # Register standard variations
            for cn in filter(None, [c_name, c_code, c_name.replace("class", "").strip()]):
                for sn in filter(None, [sec_name, sec_code, sec_name.replace("section", "").strip()]):
                    for sbn in filter(None, [sub_name, sub_code]):
                        schedule_map[(cn, sn, sbn)] = s

        # 3. Fetch all active enrolled students for these classes/sections
        class_ids = {s.class_id for s in schedules}
        stmt_st = select(Student).where(
            Student.class_id.in_(class_ids),
            Student.school_id == school_id,
            Student.tenant_id == tenant_id,
            Student.deleted_at.is_(None),
            Student.is_active == True
        )
        res_st = await self.marks_repo.db.execute(stmt_st)
        students = list(res_st.scalars().all())

        # Student lookup maps: (class_id, section_id, normalized_roll_or_adm) -> Student
        student_map = {}
        for st in students:
            if st.roll_number:
                student_map[(st.class_id, st.section_id, str(st.roll_number).strip().lower())] = st
                try:
                    student_map[(st.class_id, st.section_id, str(int(st.roll_number.strip())).lower())] = st
                except Exception:
                    pass
            if st.admission_number:
                student_map[(st.class_id, st.section_id, str(st.admission_number).strip().lower())] = st

        # 4. Parse file
        is_csv = filename.lower().endswith(".csv")
        if is_csv:
            try:
                text_content = file_bytes.decode("utf-8-sig", errors="replace")
                reader = csv.reader(io.StringIO(text_content))
                raw_rows = list(reader)
            except Exception as e:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Failed to parse CSV file: {str(e)}")
        else:
            try:
                wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
                ws = wb.active
                raw_rows = [[cell.value for cell in row] for row in ws.iter_rows()]
            except Exception as e:
                raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail=f"Failed to parse Excel workbook: {str(e)}")

        if not raw_rows or len(raw_rows) < 2:
            raise HTTPException(status_code=status.HTTP_400_BAD_REQUEST, detail="File is empty or missing data rows.")

        # Normalize header indices
        headers = [str(h or "").strip().lower() for h in raw_rows[0]]
        class_idx = next((i for i, h in enumerate(headers) if "class" in h), 0)
        sec_idx = next((i for i, h in enumerate(headers) if "section" in h or "sec" in h), 1)
        roll_idx = next((i for i, h in enumerate(headers) if "roll" in h or "admission" in h), 2)
        name_idx = next((i for i, h in enumerate(headers) if "name" in h or "student" in h), 3)
        sub_idx = next((i for i, h in enumerate(headers) if "subject" in h or "paper" in h), 4)
        max_idx = next((i for i, h in enumerate(headers) if "max" in h), -1)
        marks_idx = next((i for i, h in enumerate(headers) if "obtained" in h or "score" in h or ("mark" in h and "max" not in h)), -1)
        status_idx = next((i for i, h in enumerate(headers) if "status" in h or "result" in h or "attendance" in h), -1)
        remarks_idx = next((i for i, h in enumerate(headers) if "remark" in h or "note" in h or "comment" in h), -1)

        if marks_idx == -1 and len(headers) > 5:
            marks_idx = 5

        preview_rows: List[ExamWideUploadRowPreview] = []
        classes_detected = set()
        sections_detected = set()
        subjects_detected = set()
        students_detected = set()
        errors_summary: List[str] = []
        seen_duplicates = set()

        for row_num, row in enumerate(raw_rows[1:], start=2):
            if not row or all(v is None or str(v).strip() == "" for v in row):
                continue

            raw_class = str(row[class_idx]).strip() if class_idx < len(row) and row[class_idx] is not None else ""
            raw_sec = str(row[sec_idx]).strip() if sec_idx < len(row) and row[sec_idx] is not None else ""
            raw_roll = str(row[roll_idx]).strip() if roll_idx < len(row) and row[roll_idx] is not None else ""
            raw_name = str(row[name_idx]).strip() if name_idx < len(row) and row[name_idx] is not None else ""
            raw_sub = str(row[sub_idx]).strip() if sub_idx < len(row) and row[sub_idx] is not None else ""
            raw_max = row[max_idx] if (max_idx != -1 and max_idx < len(row)) else None
            raw_obtained = row[marks_idx] if (marks_idx != -1 and marks_idx < len(row)) else None
            raw_status_val = str(row[status_idx]).strip().upper() if (status_idx != -1 and status_idx < len(row) and row[status_idx] is not None) else "PRESENT"
            raw_remarks_val = str(row[remarks_idx]).strip() if (remarks_idx != -1 and remarks_idx < len(row) and row[remarks_idx] is not None) else ""

            row_error = None
            is_valid = True
            matched_sched = None
            matched_student = None

            # Normalization
            norm_class = raw_class.lower().replace("class", "").strip()
            norm_sec = raw_sec.lower().replace("section", "").strip()
            norm_sub = raw_sub.lower()

            if not raw_class or not raw_sec or not raw_roll or not raw_sub:
                row_error = "Missing required fields (Class, Section, Roll No, or Subject)."
                is_valid = False

            # Find matching Schedule
            if is_valid:
                matched_sched = (
                    schedule_map.get((norm_class, norm_sec, norm_sub)) or
                    schedule_map.get((raw_class.lower(), raw_sec.lower(), norm_sub)) or
                    schedule_map.get((norm_class, norm_sec, raw_sub.lower()))
                )
                if not matched_sched:
                    row_error = f"Subject '{raw_sub}' is not scheduled for Class {raw_class} Section {raw_sec} in this examination."
                    is_valid = False

            # Find matching Student
            if is_valid and matched_sched:
                norm_roll = raw_roll.lower()
                matched_student = (
                    student_map.get((matched_sched.class_id, matched_sched.section_id, norm_roll)) or
                    student_map.get((matched_sched.class_id, matched_sched.section_id, norm_roll.lstrip("0")))
                )
                if not matched_student:
                    row_error = f"Student with Roll No '{raw_roll}' not found in Class {raw_class} Section {raw_sec}."
                    is_valid = False

            # Duplicate detection in file
            if is_valid and matched_sched and matched_student:
                dup_key = (matched_student.id, matched_sched.id)
                if dup_key in seen_duplicates:
                    row_error = f"Duplicate marks entry for student '{raw_name or raw_roll}' in subject '{raw_sub}'."
                    is_valid = False
                else:
                    seen_duplicates.add(dup_key)

            # Max & Obtained Marks Validation
            max_marks_val = matched_sched.max_marks if matched_sched else 100
            if raw_max is not None:
                try:
                    max_marks_val = int(float(str(raw_max).strip()))
                except Exception:
                    pass

            marks_val = None
            result_st = "PRESENT"

            if is_valid:
                if str(raw_obtained).strip().upper() in ["AB", "ABSENT", "A"]:
                    result_st = "ABSENT"
                elif str(raw_obtained).strip().upper() in ["EX", "EXEMPTED"]:
                    result_st = "EXEMPTED"
                elif str(raw_obtained).strip().upper() in ["MP", "MALPRACTICE"]:
                    result_st = "MALPRACTICE"
                elif raw_obtained is not None and str(raw_obtained).strip() != "":
                    try:
                        marks_val = round(float(str(raw_obtained).strip()), 2)
                        if marks_val < 0:
                            row_error = f"Marks cannot be negative ({marks_val})."
                            is_valid = False
                        elif marks_val > max_marks_val:
                            row_error = f"Marks obtained ({marks_val}) exceeds Maximum Marks ({max_marks_val})."
                            is_valid = False
                    except ValueError:
                        row_error = f"Invalid marks numeric value '{raw_obtained}'."
                        is_valid = False
                else:
                    marks_val = None

            if raw_class: classes_detected.add(raw_class)
            if raw_sec: sections_detected.add(f"{raw_class}-{raw_sec}")
            if raw_sub: subjects_detected.add(raw_sub)
            if matched_student: students_detected.add(str(matched_student.id))

            if not is_valid and row_error:
                errors_summary.append(f"Row {row_num} (Class {raw_class}-{raw_sec}, {raw_sub}, Roll {raw_roll}): {row_error}")

            preview_rows.append(
                ExamWideUploadRowPreview(
                    row_number=row_num,
                    class_name=raw_class,
                    section_name=raw_sec,
                    roll_number=raw_roll,
                    student_name=raw_name or (matched_student.full_name if matched_student else None),
                    subject_name=raw_sub,
                    max_marks=max_marks_val,
                    marks_obtained=marks_val,
                    status=result_st,
                    remarks=raw_remarks_val,
                    is_valid=is_valid,
                    error_message=row_error,
                    student_id=str(matched_student.id) if matched_student else None,
                    exam_schedule_id=str(matched_sched.id) if matched_sched else None,
                    class_id=str(matched_sched.class_id) if matched_sched else None,
                    section_id=str(matched_sched.section_id) if matched_sched else None,
                    subject_id=str(matched_sched.subject_id) if matched_sched else None
                )
            )

        valid_count = sum(1 for r in preview_rows if r.is_valid)
        invalid_count = len(preview_rows) - valid_count

        return ExamWideUploadPreviewResponse(
            total_rows=len(preview_rows),
            valid_rows_count=valid_count,
            invalid_rows_count=invalid_count,
            classes_detected=sorted(list(classes_detected)),
            sections_detected=sorted(list(sections_detected)),
            subjects_detected=sorted(list(subjects_detected)),
            students_count=len(students_detected),
            errors=errors_summary,
            preview_rows=preview_rows
        )

    async def confirm_exam_wide_marks(
        self,
        tenant_id: uuid.UUID,
        school_id: uuid.UUID,
        req: ExamWideUploadConfirmRequest,
        current_user: User
    ) -> ExamWideUploadSummary:
        # Load Examination
        stmt_e = select(Examination).where(
            Examination.id == req.exam_id,
            Examination.school_id == school_id,
            Examination.tenant_id == tenant_id,
            Examination.deleted_at.is_(None)
        )
        res_e = await self.marks_repo.db.execute(stmt_e)
        examination = res_e.scalar_one_or_none()
        if not examination:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Examination not found.")

        # Fetch existing marks for this exam to update existing records
        stmt_existing = select(Marks).where(
            Marks.examination_id == req.exam_id,
            Marks.school_id == school_id,
            Marks.tenant_id == tenant_id,
            Marks.deleted_at.is_(None)
        )
        res_ex = await self.marks_repo.db.execute(stmt_existing)
        existing_marks_list = list(res_ex.scalars().all())
        existing_map = {(m.student_id, m.exam_schedule_id): m for m in existing_marks_list}

        # Fetch schedules to get academic_year_id, teacher_id, etc.
        stmt_s = select(ExamSchedule).where(
            ExamSchedule.exam_id == req.exam_id,
            ExamSchedule.school_id == school_id,
            ExamSchedule.tenant_id == tenant_id,
            ExamSchedule.deleted_at.is_(None)
        )
        res_s = await self.marks_repo.db.execute(stmt_s)
        schedules = {s.id: s for s in res_s.scalars().all()}

        # Fetch TeacherSubjectAssignments to resolve teacher_id
        stmt_tsa = select(TeacherSubjectAssignment).where(
            TeacherSubjectAssignment.school_id == school_id,
            TeacherSubjectAssignment.tenant_id == tenant_id,
            TeacherSubjectAssignment.deleted_at.is_(None)
        )
        res_tsa = await self.marks_repo.db.execute(stmt_tsa)
        tsa_map = {tsa.id: (tsa.teacher_id, tsa.id) for tsa in res_tsa.scalars().all()}

        fallback_teacher_id = None
        teacher_stmt = select(Teacher.id).where(
            Teacher.school_id == school_id,
            Teacher.tenant_id == tenant_id,
            Teacher.deleted_at.is_(None)
        ).limit(1)
        fallback_teacher_id = (await self.marks_repo.db.execute(teacher_stmt)).scalar()

        saved_count = 0
        failed_count = 0
        students_set = set()
        classes_set = set()
        sections_set = set()
        subjects_set = set()

        target_status = MarksStatus.PUBLISHED if req.auto_approve else MarksStatus.SUBMITTED

        for row in req.rows:
            if not row.is_valid or not row.student_id or not row.exam_schedule_id:
                failed_count += 1
                continue

            try:
                st_id = uuid.UUID(row.student_id)
                sched_id = uuid.UUID(row.exam_schedule_id)
                sched = schedules.get(sched_id)
                if not sched:
                    failed_count += 1
                    continue

                res_status_enum = ExamResult(row.status) if row.status in ExamResult.__members__ else ExamResult.PRESENT
                grade_val = self._compute_grade(row.marks_obtained, row.max_marks) if row.marks_obtained is not None else None

                tsa_info = tsa_map.get(sched.teacher_subject_assignment_id) if sched.teacher_subject_assignment_id else None
                t_id = tsa_info[0] if tsa_info else fallback_teacher_id
                tsa_id = sched.teacher_subject_assignment_id or (tsa_info[1] if tsa_info else None)

                existing = existing_map.get((st_id, sched_id))
                if existing:
                    existing.marks_obtained = row.marks_obtained
                    existing.maximum_marks = row.max_marks
                    existing.result_status = res_status_enum
                    existing.status = target_status
                    existing.grade = grade_val
                    existing.remarks = row.remarks
                    existing.updated_by = current_user.id
                    self.marks_repo.db.add(existing)
                else:
                    new_mark = Marks(
                        tenant_id=tenant_id,
                        school_id=school_id,
                        academic_year_id=sched.academic_year_id,
                        examination_id=req.exam_id,
                        exam_schedule_id=sched_id,
                        student_id=st_id,
                        teacher_subject_assignment_id=tsa_id,
                        teacher_id=t_id,
                        subject_id=sched.subject_id,
                        class_id=sched.class_id,
                        section_id=sched.section_id,
                        maximum_marks=row.max_marks,
                        marks_obtained=row.marks_obtained,
                        result_status=res_status_enum,
                        status=target_status,
                        grade=grade_val,
                        remarks=row.remarks,
                        created_by=current_user.id,
                        updated_by=current_user.id
                    )
                    self.marks_repo.db.add(new_mark)

                saved_count += 1
                students_set.add(row.student_id)
                classes_set.add(row.class_name)
                sections_set.add(f"{row.class_name}-{row.section_name}")
                subjects_set.add(row.subject_name)
            except Exception as e:
                logger.error(f"Error saving row {row.row_number}: {e}")
                failed_count += 1

        await self.marks_repo.db.commit()

        return ExamWideUploadSummary(
            examination_id=req.exam_id,
            examination_name=examination.exam_name,
            students_processed=len(students_set),
            classes_count=len(classes_set),
            sections_count=len(sections_set),
            subjects_count=len(subjects_set),
            total_records=len(req.rows),
            saved_count=saved_count,
            failed_count=failed_count
        )

    async def publish_examination_marks(
        self,
        tenant_id: uuid.UUID,
        school_id: uuid.UUID,
        exam_id: uuid.UUID,
        current_user: User
    ) -> ExaminationPublishSummary:
        # 1. Fetch Examination
        stmt_e = select(Examination).where(
            Examination.id == exam_id,
            Examination.school_id == school_id,
            Examination.tenant_id == tenant_id,
            Examination.deleted_at.is_(None)
        )
        res_e = await self.marks_repo.db.execute(stmt_e)
        examination = res_e.scalar_one_or_none()
        if not examination:
            raise HTTPException(status_code=status.HTTP_404_NOT_FOUND, detail="Examination not found.")

        # 2. Fetch all schedules
        stmt_s = select(ExamSchedule).where(
            ExamSchedule.exam_id == exam_id,
            ExamSchedule.school_id == school_id,
            ExamSchedule.tenant_id == tenant_id,
            ExamSchedule.deleted_at.is_(None)
        ).options(
            joinedload(ExamSchedule.class_obj),
            joinedload(ExamSchedule.section),
            joinedload(ExamSchedule.subject)
        )
        res_s = await self.marks_repo.db.execute(stmt_s)
        schedules = list(res_s.unique().scalars().all())
        if not schedules:
            raise HTTPException(status_code=status.HTTP_422_UNPROCESSABLE_ENTITY, detail="No exam schedules found for this examination.")

        # 3. Calculate expected vs entered
        total_expected = 0
        marks_entered_count = 0
        missing_count = 0
        missing_breakdown = []

        # Fetch all marks for this examination
        stmt_m = select(Marks).where(
            Marks.examination_id == exam_id,
            Marks.school_id == school_id,
            Marks.tenant_id == tenant_id,
            Marks.deleted_at.is_(None)
        )
        res_m = await self.marks_repo.db.execute(stmt_m)
        all_marks = list(res_m.scalars().all())
        marks_by_sched = {}
        for m in all_marks:
            marks_by_sched.setdefault(m.exam_schedule_id, []).append(m)

        for s in schedules:
            students = await self.marks_repo.get_class_students_sorted(s.class_id, s.section_id, school_id, tenant_id)
            sched_expected = len(students)
            sched_marks = marks_by_sched.get(s.id, [])
            sched_entered = len(sched_marks)
            sched_missing = max(0, sched_expected - sched_entered)

            total_expected += sched_expected
            marks_entered_count += sched_entered
            missing_count += sched_missing

            if sched_missing > 0:
                missing_breakdown.append({
                    "class_name": s.class_obj.name if s.class_obj else "Class",
                    "section_name": s.section.name if s.section else "Section",
                    "subject_name": s.subject.subject_name if s.subject else "Subject",
                    "class_id": str(s.class_id),
                    "section_id": str(s.section_id),
                    "schedule_id": str(s.id),
                    "missing_count": sched_missing,
                    "expected_count": sched_expected,
                    "entered_count": sched_entered
                })

        # 4. Transition all marks to PUBLISHED
        published_count = 0
        now_iso = datetime.now(timezone.utc).isoformat()
        for m in all_marks:
            if m.status != MarksStatus.LOCKED:
                m.status = MarksStatus.PUBLISHED
                m.audit_history = list(m.audit_history or []) + [{
                    "action": "EXAM_WIDE_PUBLISH",
                    "reason": "Complete examination publication",
                    "updated_by": str(current_user.id),
                    "updated_at": now_iso
                }]
                m.updated_by = current_user.id
                self.marks_repo.db.add(m)
                published_count += 1

        # Also update Examination status to PUBLISHED if it was ongoing / marks entry / under review
        if examination.status not in [ExamStatus.COMPLETED, ExamStatus.ARCHIVED]:
            examination.status = ExamStatus.PUBLISHED
            self.marks_repo.db.add(examination)

        await self.marks_repo.db.commit()

        return ExaminationPublishSummary(
            examination_id=exam_id,
            examination_name=examination.exam_name,
            total_expected_records=total_expected,
            marks_entered_count=marks_entered_count,
            published_count=published_count,
            missing_count=missing_count,
            is_fully_published=(missing_count == 0),
            missing_breakdown=missing_breakdown
        )

    async def generate_exam_wide_template(
        self, tenant_id: uuid.UUID, school_id: uuid.UUID, exam_id: uuid.UUID
    ) -> bytes:
        import io
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment

        # Fetch schedules & enrolled students
        stmt_s = select(ExamSchedule).where(
            ExamSchedule.exam_id == exam_id,
            ExamSchedule.school_id == school_id,
            ExamSchedule.tenant_id == tenant_id,
            ExamSchedule.deleted_at.is_(None)
        ).options(
            joinedload(ExamSchedule.class_obj),
            joinedload(ExamSchedule.section),
            joinedload(ExamSchedule.subject)
        ).order_by(ExamSchedule.class_id, ExamSchedule.section_id)

        res_s = await self.marks_repo.db.execute(stmt_s)
        schedules = list(res_s.unique().scalars().all())

        wb = Workbook()
        ws = wb.active
        ws.title = "Exam Marks"

        # Headers
        headers = ["Class", "Section", "Roll No", "Student Name", "Subject", "Max Marks", "Obtained Marks", "Status", "Remarks"]
        ws.append(headers)

        header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="1A365D", end_color="1A365D", fill_type="solid")
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Populate rows for each schedule & student
        for s in schedules:
            students = await self.marks_repo.get_class_students_sorted(s.class_id, s.section_id, school_id, tenant_id)
            c_name = s.class_obj.name if s.class_obj else "Class"
            sec_name = s.section.name if s.section else "Section"
            sub_name = s.subject.subject_name if s.subject else "Subject"

            for st in students:
                ws.append([
                    c_name,
                    sec_name,
                    st.roll_number or "",
                    st.full_name or "",
                    sub_name,
                    s.max_marks,
                    "",
                    "PRESENT",
                    ""
                ])

        # Auto-adjust column widths
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            col_letter = col[0].column_letter
            ws.column_dimensions[col_letter].width = max(max_len + 4, 12)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()


