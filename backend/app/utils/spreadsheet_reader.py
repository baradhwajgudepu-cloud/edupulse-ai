import csv
import io
import re
import datetime
import zipfile
from typing import List, Tuple, Optional

def normalize_header(header: str) -> str:
    """
    Standardize spacing, casing, underscores, and hyphens to lowercase with single underscores.
    e.g., 'Teacher ID' -> 'teacher_id', 'Teacher-ID' -> 'teacher_id', ' teacher id ' -> 'teacher_id'.
    """
    if not header:
        return ""
    # Strip spaces and convert to lowercase
    h = header.strip().lower()
    # Replace any sequence of whitespace, dashes, or underscores with a single underscore
    h = re.sub(r'[\s\-_]+', '_', h)
    return h

def format_cell_value(val) -> str:
    """
    Converts spreadsheet cell values (numbers, dates, booleans) into standard formatted strings.
    Prevents float representation (e.g. '101.0') for whole numbers.
    """
    if val is None:
        return ""
    if isinstance(val, bool):
        return "true" if val else "false"
    if isinstance(val, (int, float)):
        # Check if float represents a whole number
        if isinstance(val, float) and val.is_integer():
            return str(int(val))
        return str(val)
    if isinstance(val, (datetime.datetime, datetime.date)):
        return val.strftime("%Y-%m-%d")
    return str(val).strip()

def detect_file_type(content: bytes, filename: str = None) -> str:
    """
    Detects the file type by checking magic number bytes signatures or file extension.
    Returns: 'csv', 'xlsx', 'xls', 'xlsb', 'ods'.
    """
    if not content:
        return 'csv'

    # Check old XLS binary format signature (OLECF)
    if content.startswith(b"\xD0\xCF\x11\xE0\xA1\xB1\x1A\xE1"):
        return 'xls'

    # Check ZIP-based format signatures (OpenXML / ODF)
    if content.startswith(b"PK\x03\x04"):
        try:
            with zipfile.ZipFile(io.BytesIO(content)) as z:
                names = z.namelist()
                if "content.xml" in names:
                    return 'ods'
                elif any("workbook.bin" in name for name in names):
                    return 'xlsb'
                else:
                    return 'xlsx'  # default fallback for standard xlsx / xlsm zip
        except Exception:
            pass

    # Fallback to filename extension
    if filename:
        ext = filename.split(".")[-1].lower()
        if ext in ['csv', 'xlsx', 'xls', 'xlsm', 'xlsb', 'ods']:
            # xlsx and xlsm use the same openpyxl reader
            return 'xlsx' if ext == 'xlsm' else ext

        # If it parses as text, treat as CSV
    if b"\x00" in content[:4096]:
        raise ValueError("Unsupported or unrecognized spreadsheet file format.")

    try:
        content[:1024].decode('utf-8-sig')
        return 'csv'
    except UnicodeDecodeError:
        pass

    raise ValueError("Unsupported or unrecognized spreadsheet file format.")

def read_spreadsheet(
    content: bytes,
    filename: str = None,
    sheet_name: str = None
) -> Tuple[List[str], str, List[List[str]]]:
    """
    Reads the content of any supported spreadsheet and returns:
    (list_of_sheet_names, selected_sheet_name, raw_data_rows)
    """
    file_type = detect_file_type(content, filename)

    if file_type == 'csv':
        # CSV files only have one implicit sheet named "CSV"
        default_sheet = "CSV"
        text_content = content.decode("utf-8-sig")
        reader = csv.reader(io.StringIO(text_content))
        rows = [row for row in reader]
        # Clean empty cells
        cleaned_rows = []
        for r in rows:
            row_vals = [str(c).strip() for c in r]
            if any(row_vals):
                cleaned_rows.append(row_vals)
        return [default_sheet], default_sheet, cleaned_rows

    elif file_type == 'xlsx':
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
        sheets = wb.sheetnames
        if not sheets:
            raise ValueError("Spreadsheet workbook contains no sheets.")
        
        target_sheet = sheet_name if (sheet_name and sheet_name in sheets) else sheets[0]
        ws = wb[target_sheet]

        rows = []
        for r in ws.iter_rows(values_only=True):
            row_vals = [format_cell_value(c) for c in r]
            # Strip trailing empty cells
            while row_vals and row_vals[-1] == "":
                row_vals.pop()
            if any(row_vals):
                rows.append(row_vals)
        wb.close()
        return sheets, target_sheet, rows

    elif file_type == 'xls':
        import xlrd
        wb = xlrd.open_workbook(file_contents=content)
        sheets = wb.sheet_names()
        if not sheets:
            raise ValueError("Spreadsheet workbook contains no sheets.")

        target_sheet = sheet_name if (sheet_name and sheet_name in sheets) else sheets[0]
        ws = wb.sheet_by_name(target_sheet)

        rows = []
        for r_idx in range(ws.nrows):
            row_vals = []
            for c_idx in range(ws.ncols):
                cell = ws.cell(r_idx, c_idx)
                val = cell.value
                # If date type, convert using xlrd helper
                if cell.ctype == xlrd.XL_CELL_DATE:
                    dt_tuple = xlrd.xldate_as_tuple(val, wb.datemode)
                    val = datetime.date(*dt_tuple[:3])
                row_vals.append(format_cell_value(val))

            while row_vals and row_vals[-1] == "":
                row_vals.pop()
            if any(row_vals):
                rows.append(row_vals)
        return sheets, target_sheet, rows

    elif file_type == 'xlsb':
        from pyxlsb import open_workbook
        with open_workbook(io.BytesIO(content)) as wb:
            sheets = wb.sheets
            if not sheets:
                raise ValueError("Spreadsheet workbook contains no sheets.")

            target_sheet = sheet_name if (sheet_name and sheet_name in sheets) else sheets[0]
            rows = []
            with wb.get_sheet(target_sheet) as ws:
                for r in ws.rows():
                    row_vals = [format_cell_value(c.v) for c in r]
                    while row_vals and row_vals[-1] == "":
                        row_vals.pop()
                    if any(row_vals):
                        rows.append(row_vals)
            return sheets, target_sheet, rows

    elif file_type == 'ods':
        from odf.opendocument import load
        from odf.table import Table, TableRow, TableCell
        from odf.text import P

        doc = load(io.BytesIO(content))
        tables = doc.spreadsheet.getElementsByType(Table)
        if not tables:
            raise ValueError("Spreadsheet workbook contains no sheets.")

        sheets = [t.getAttribute("name") for t in tables]
        target_sheet = sheet_name if (sheet_name and sheet_name in sheets) else sheets[0]
        
        # Locate target table
        table = None
        for t in tables:
            if t.getAttribute("name") == target_sheet:
                table = t
                break
        if not table:
            table = tables[0]

        rows = []
        table_rows = table.getElementsByType(TableRow)
        for row in table_rows:
            cells = row.getElementsByType(TableCell)
            row_vals = []
            for cell in cells:
                repeated = cell.getAttribute("numbercolumnsrepeated")
                repeat_count = int(repeated) if repeated else 1

                paragraphs = cell.getElementsByType(P)
                if paragraphs:
                    text = "".join(str(node) for p in paragraphs for node in p.childNodes if node.nodeType == 3)
                    val = text.strip()
                else:
                    val = ""

                # Avoid huge repetition count for trailing empty columns
                if val == "" and repeat_count > 1:
                    repeat_count = 1

                for _ in range(repeat_count):
                    row_vals.append(val)

            # Strip trailing empty cells
            while row_vals and row_vals[-1] == "":
                row_vals.pop()

            if any(row_vals):
                rows.append(row_vals)

        return sheets, target_sheet, rows

    raise ValueError(f"Unsupported spreadsheet format: {file_type}")
